from pickletools import optimize
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from datetime import datetime
import os.path
import os

column_widths = []
print("Loading files")
allsku = openpyxl.load_workbook('All_SKUs_21.06.xlsx')
allsku = allsku[allsku.sheetnames[0]]
print("Loaded all sku")

glossary = openpyxl.load_workbook('glossary.xlsx')
glossary = glossary[glossary.sheetnames[2]]
glossary.delete_rows(1, 4)
print("Loaded glossary")


headers = [cell.value for cell in allsku[1]]  # all product column names

print(headers)
#           category name   mandatory


categories = {}
for row in glossary:
    categoryid = str(row[0].value)
    if not categories.get(categoryid, None):
        categories[categoryid] = [(row[5].value, row[2].value)]
    else:
        categories[categoryid].extend([(row[5].value, row[2].value)])

for i, row in enumerate(allsku):
    if i == 0:  # skip headers
        continue
    gs1 = str(row[0].value)
    categoryid = str(row[4].value)
    if categoryid not in categories.keys():
        categoryid = '10000334'  # UNKNOWN CAT ALL TREATED AS : Bath/shower accessory type

    if not os.path.exists(categoryid):
        os.mkdir(categoryid)

    clean_wb = Workbook()
    clean_ws = clean_wb.active

    local_req = headers[::]
    req = [column for column, priority in categories[categoryid] if priority == 1]
    local_req = req[::]
    req_vals = []
    for i, cell in enumerate(row):
        for required in local_req:
            if headers[i].lower().startswith(required.lower()):
                to_remove = required
                break
        else:
            continue
        local_req.remove(to_remove)
        req_vals.append(cell.value)

    # req_vals = [cell.value for i, cell in enumerate(row) if any([required.lower().startswith(headers[i].lower()) for required in req])]

    opt = [column for column, priority in categories[categoryid] if priority == 0]
    opt_vals = list()
    for i, cell in enumerate(row):
        for option in opt:
            if headers[i].lower().startswith(option.lower()):
                opt_vals.append(cell.value)

    print(len(req))
    print(len(req_vals))
    print()
    clean_wb = Workbook()
    clean_ws = clean_wb.active
    clean_ws.append(["GS1 item number (GTIN):", gs1, "BRICK CAT:", categoryid, "v0.1",
                     "Script last executed:", datetime.today().strftime('%Y-%m-%d %H:%M:%S')])
    clean_ws.append(req)
    clean_ws.append(req_vals)
    clean_ws.append(opt)
    clean_ws.append(opt_vals)
    for cell in clean_ws['A2:'+chr(len(req_vals)+65)+'2'][0]:
        cell.fill = PatternFill("solid", start_color="ed5587")
        cell.font = Font(bold=True)
    for cell in clean_ws['A4:BR4'][0]:
        cell.fill = PatternFill("solid", start_color="97ed55")
        cell.font = Font(italic=True)
    for row in clean_ws:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell)) > column_widths[i]:
                    column_widths[i] = len(cell)
            else:
                column_widths += [len(str(cell))]
    for i, column_width in enumerate(column_widths, 1):
        clean_ws.column_dimensions[get_column_letter(
            i)].width = column_width
    clean_wb.save(str(categoryid)+"\\"+str(gs1)+".xlsx")
    req = list()
    req_vals = list()
    opt = list()
    opt_vals = list()
    clean_wb.close()
    # print(opt)
    # print(opt_vals)
